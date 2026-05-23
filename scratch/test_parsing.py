import openpyxl

file_path = "/Users/alejandracastro/Desktop/ALTITUD HUB/Performance Dashboard -ALTITUD.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Production']

# 1. Parse Commissions Section (Rows 4-19)
print("=== SECTION 1: COMMISSIONS ===")
years = []
for c in range(2, 4): # Cols B and C
    val = sheet.cell(5, c).value
    if val:
        years.append((c, int(val)))

for col_idx, year in years:
    print(f"Year: {year}")
    for r in range(6, 18): # Rows 6-17 (Jan-Dec)
        month = sheet.cell(r, 1).value
        val = sheet.cell(r, col_idx).value
        print(f"  {month}: {val}")

# 2. Parse Agent Count Section (Rows 22-36)
print("\n=== SECTION 2: AGENT COUNT ===")
ac_years = []
for c in range(2, 4):
    val = sheet.cell(23, c).value
    if val:
        ac_years.append((c, int(val)))

for col_idx, year in ac_years:
    print(f"Year: {year}")
    for r in range(24, 36):
        month = sheet.cell(r, 1).value
        val = sheet.cell(r, col_idx).value
        print(f"  {month}: {val}")

# 3. Parse Monthly Commissions Per Associate (Rows 39-53)
print("\n=== SECTION 3: MONTHLY COMMISSIONS PER ASSOCIATE (2026) ===")
months_header = []
c = 2
while True:
    val = sheet.cell(40, c).value
    if not val or val == 'Total':
        break
    months_header.append((c, val))
    c += 1
print(f"Months with data: {[m[1] for m in months_header]}")

for r in range(41, 53):
    agent_name = sheet.cell(r, 1).value
    print(f"Agent: {agent_name}")
    for col_idx, month in months_header:
        val = sheet.cell(r, col_idx).value
        print(f"  {month}: {val}")

# 4. Parse Monthly Volume Per Associate (Rows 62-76)
print("\n=== SECTION 4: MONTHLY VOLUME PER ASSOCIATE (2026) ===")
volume_months_header = []
c = 2
while True:
    val = sheet.cell(63, c).value
    if not val or val == 'Total':
        break
    volume_months_header.append((c, val))
    c += 1

for r in range(64, 76):
    agent_name = sheet.cell(r, 1).value
    print(f"Agent: {agent_name}")
    for col_idx, month in volume_months_header:
        val = sheet.cell(r, col_idx).value
        print(f"  {month}: {val}")
