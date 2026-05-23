import openpyxl

file_path = "/Users/alejandracastro/Desktop/ALTITUD HUB/Performance Dashboard -ALTITUD.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Production']

print(f"Total rows in sheet: {sheet.max_row}")
for r in range(77, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    while row_vals and row_vals[-1] is None:
        row_vals.pop()
    if any(v is not None for v in row_vals):
        formatted_vals = [f"'{v}'" if isinstance(v, str) else str(v) for v in row_vals]
        print(f"Row {r:02d}: {formatted_vals}")
